from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


def generate_boq_excel(boq_data, grand_total, project_name="My Project",
                       building_type="G+2", col_data=None, beam_data=None):

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.sheet_view.showGridLines = False

    # ── Styles ──────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="30363d")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center")

    def right():
        return Alignment(horizontal="right", vertical="center")

    # ── Column widths ────────────────────────────────────────────
    col_widths = [5, 8, 45, 8, 14, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"].value = f"BILL OF QUANTITIES"
    ws["A1"].font = Font(bold=True, size=14, color="00c896")
    ws["A1"].fill = fill("0d1117")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Project: {project_name}  |  Type: {building_type}  |  Date: {datetime.now().strftime('%d %B %Y')}"
    ws["A2"].font = Font(size=10, color="8b949e")
    ws["A2"].fill = fill("0d1117")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:G3")
    ws["A3"].fill = fill("0d1117")
    ws.row_dimensions[3].height = 8

    # ── Table header ─────────────────────────────────────────────
    headers = ["#", "Type", "Description", "Unit", "Quantity", "Rate (₹)", "Amount (₹)"]
    for col_num, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_num, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = fill("161b22")
        c.alignment = center()
        c.border = border()
    ws.row_dimensions[4].height = 22

    # ── Write row helper ─────────────────────────────────────────
    def write_row(row, serial, tag, desc, unit, qty, rate, amount, bg):
        data = [serial, tag, desc, unit, round(qty, 3), round(rate, 2), round(amount, 2)]
        for col_num, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col_num, value=val)
            c.fill = fill(bg)
            c.font = Font(size=10, color="e6edf3")
            c.border = border()
            if col_num in (5, 6, 7):
                c.alignment = right()
                c.number_format = '#,##0.00'
            else:
                c.alignment = left()
        ws.row_dimensions[row].height = 18

    # ── Section header helper ────────────────────────────────────
    def write_section(row, label, bg):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = f"   {label}"
        c.font = Font(bold=True, size=10, color="00c896")
        c.fill = fill(bg)
        c.alignment = left()
        c.border = border()
        ws.row_dimensions[row].height = 20

    current_row = 5
    serial = 1

    # ── FOOTING ──────────────────────────────────────────────────
    write_section(current_row, "FOUNDATION — ISOLATED FOOTINGS", "1c2230")
    current_row += 1

    footing_count = col_data.get("footing_count", 1) if col_data else 1

    for item in boq_data:
        if item.get("item_no", 0) <= 4:
            write_row(current_row, serial, "FTG",
                      item["description"], item["unit"],
                      round(item["quantity"] * footing_count, 3),
                      item["rate"],
                      round(item["amount"] * footing_count, 2), "1c2230")
            current_row += 1
            serial += 1 
    

    # ── COLUMN ───────────────────────────────────────────────────
    if col_data:
        write_section(current_row, "COLUMNS — RCC", "1a2035")
        current_row += 1

        col_items = [
            ("RCC M25 Column Concrete", "m³",
             col_data.get("colRCC", 0), col_data.get("cRccRate", 0),
             col_data.get("colRCC", 0) * col_data.get("cRccRate", 0)),

            ("Column Steel Fe415 (Main Bars + Ties)", "kg",
             col_data.get("colSteel", 0), col_data.get("cSteelRate", 0),
             col_data.get("colSteel", 0) * col_data.get("cSteelRate", 0)),

            ("Column 4-sided Formwork", "m²",
             col_data.get("colFW", 0), col_data.get("cFwRate", 0),
             col_data.get("colFW", 0) * col_data.get("cFwRate", 0)),
        ]

        for desc, unit, qty, rate, amt in col_items:
            write_row(current_row, serial, "COL", desc, unit, qty, rate, amt, "1a2035")
            current_row += 1
            serial += 1

    # ── BEAM ─────────────────────────────────────────────────────
    if beam_data:
        write_section(current_row, "BEAMS — RCC", "1c2035")
        current_row += 1

        beam_items = [
            ("RCC M25 Beam Concrete", "m³",
             beam_data.get("beamRCC", 0), beam_data.get("bRccRate", 0),
             beam_data.get("beamRCC", 0) * beam_data.get("bRccRate", 0)),

            ("Beam Steel Fe415 (Main Bars + Stirrups)", "kg",
             beam_data.get("beamSteel", 0), beam_data.get("bSteelRate", 0),
             beam_data.get("beamSteel", 0) * beam_data.get("bSteelRate", 0)),

            ("Beam Side + Soffit Formwork", "m²",
             beam_data.get("beamFW", 0), beam_data.get("bFwRate", 0),
             beam_data.get("beamFW", 0) * beam_data.get("bFwRate", 0)),
        ]

        for desc, unit, qty, rate, amt in beam_items:
            write_row(current_row, serial, "BM", desc, unit, qty, rate, amt, "1c2035")
            current_row += 1
            serial += 1

    # ── SLAB ─────────────────────────────────────────────────────
    write_section(current_row, "SLAB — RCC", "1c2128")
    current_row += 1

    for item in boq_data:
        if item.get("item_no", 0) > 4:
            write_row(current_row, serial, "SLB",
                      item["description"], item["unit"],
                      item["quantity"], item["rate"], item["amount"], "1c2128")
            current_row += 1
            serial += 1

    # ── GRAND TOTAL ───────────────────────────────────────────────
    current_row += 1
    ws.merge_cells(f"A{current_row}:F{current_row}")
    label_cell = ws[f"A{current_row}"]
    label_cell.value = "GRAND TOTAL"
    label_cell.font = Font(bold=True, size=13, color="0d1117")
    label_cell.fill = fill("00c896")
    label_cell.alignment = right()
    label_cell.border = border()

    total_cell = ws.cell(row=current_row, column=7, value=round(grand_total, 2))
    total_cell.font = Font(bold=True, size=13, color="0d1117")
    total_cell.fill = fill("00c896")
    total_cell.alignment = right()
    total_cell.border = border()
    total_cell.number_format = '#,##0.00'
    ws.row_dimensions[current_row].height = 28

    # ── Save ──────────────────────────────────────────────────────
    download_folder = "downloads"
    os.makedirs(download_folder, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"BOQ_{project_name.replace(' ', '_')}_{timestamp}.xlsx"
    file_path = os.path.join(download_folder, file_name)
    wb.save(file_path)
    return file_path